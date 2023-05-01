from openpyxl import load_workbook, Workbook
import os


def get_workbook(wb_dir, headings):

    workbook = None

    if os.path.isfile(wb_dir):

        workbook = load_workbook(wb_dir)

    else:

        workbook = Workbook()

        page = workbook.active

        page.append(headings)

    return workbook


def extract_employee(txt_data):

    temp_emp = txt_data.split('-')

    return {
        'name': temp_emp[0].strip(),
        'role': temp_emp[1].strip(),
        'company': temp_emp[2].strip(),
    }


def get_employees(data):

    employees = []

    for d in data:

        employee = extract_employee(d)

        employees.append(employee)

    return employees


def add_data_to_excel(emp_data):
    excel_headings = ['اسم الموظف', 'المنصب', 'اسم الشركة']

    workbook = get_workbook('potential-employees.xlsx', excel_headings)

    worksheet = workbook.active

    emp = [
        emp_data['name'],
        emp_data['role'],
        emp_data['company'],
    ]

    worksheet.append(emp)

    workbook.save('potential-employees.xlsx')


if __name__ == '__main__':

    with open('data.txt', 'r', encoding='utf-8') as f:

        txt = f.read()

    txt_lines = txt.split('\n')

    employees_data = get_employees(txt_lines)

    for ed in employees_data:
        add_data_to_excel(ed)
